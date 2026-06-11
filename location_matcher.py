import os
import sys
import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
from google import genai
from google.genai import types
from pydantic import BaseModel, Field

# Load environment variables
load_dotenv('backend/.env')
load_dotenv('.env')

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if not GEMINI_API_KEY:
    print("Error: GEMINI_API_KEY not found in environment or backend/.env file.")
    sys.exit(1)

# Initialize GenAI Client
client = genai.Client(api_key=GEMINI_API_KEY)

class LocationMatchResult(BaseModel):
    work_city: str | None = Field(description="The standardized, canonical city name of the Work Location (e.g. 'New York'). Use Title Case. Null if not specified or unclear.")
    work_country: str | None = Field(description="The standardized, canonical country name of the Work Location (e.g. 'United States'). Use Title Case. Null if not specified or unclear.")
    updated_city: str | None = Field(description="The standardized, canonical city name of the Updated Location (e.g. 'New York'). Use Title Case. Null if not specified or unclear.")
    updated_country: str | None = Field(description="The standardized, canonical country name of the Updated Location (e.g. 'United States'). Use Title Case. Null if not specified or unclear.")

def standardize_and_match(input_file="input.xlsx", output_file="output.xlsx"):
    print(f"Reading input file: {input_file}...")
    try:
        df = pd.read_excel(input_file)
    except Exception as e:
        print(f"Error reading file {input_file}: {e}")
        sys.exit(1)

    # Detect column headers case-insensitively
    work_col = next((c for c in df.columns if 'work' in c.lower() and 'location' in c.lower()), None)
    updated_col = next((c for c in df.columns if 'updated' in c.lower() and 'location' in c.lower()), None)

    if not work_col:
        print("Warning: Could not automatically detect a 'Work Location' column. Using default 'Work Location'.")
        work_col = 'Work Location'
        if work_col not in df.columns:
            df[work_col] = ""

    if not updated_col:
        print("Warning: Could not automatically detect an 'Updated Location' column. Using default 'Updated Location'.")
        updated_col = 'Updated Location'
        if updated_col not in df.columns:
            df[updated_col] = ""

    print(f"Detected columns -> Work Location: '{work_col}', Updated Location: '{updated_col}'")

    # Add output columns if they don't exist
    new_cols = ['Work City', 'Work Country', 'Updated City', 'Updated Country']
    for col in new_cols:
        df[col] = ""

    # Process each row
    print("Processing location data with Gemini...")
    for idx, row in df.iterrows():
        work_loc = str(row[work_col]).strip() if pd.notna(row[work_col]) else ""
        updated_loc = str(row[updated_col]).strip() if pd.notna(row[updated_col]) else ""

        print(f"Row {idx + 1}: Work='{work_loc}', Updated='{updated_loc}'")

        if not work_loc and not updated_loc:
            continue

        prompt = f"""
        You are an expert geographical location parser. Standardize the following location strings into canonical City and Country names.
        
        Work Location String: "{work_loc}"
        Updated Location String: "{updated_loc}"
        
Rule:
        1. Parse and extract the City and Country.
        2. Standardize names to their canonical forms (e.g., "NYC" or "New York City" should become "New York"; "USA" or "United States of America" should become "United States"; "UK" or "United Kingdom of Great Britain" should become "United Kingdom"; "Bangalore" or "Bengaluru" should become "Bangalore").
        3. If a field is missing or cannot be resolved, return null.
        """

        import time
        max_retries = 5
        retry_delay = 2
        success = False
        response = None
        
        for attempt in range(max_retries):
            try:
                response = client.models.generate_content(
                    model='gemini-2.5-flash',
                    contents=prompt,
                    config=types.GenerateContentConfig(
                        response_mime_type="application/json",
                        response_schema=LocationMatchResult,
                    ),
                )
                success = True
                break
            except Exception as e:
                err_msg = str(e).lower()
                if "503" in err_msg or "429" in err_msg or "unavailable" in err_msg or "rate limit" in err_msg:
                    print(f"  Attempt {attempt + 1} failed: {e}. Retrying in {retry_delay}s...")
                    time.sleep(retry_delay)
                    retry_delay *= 2
                else:
                    print(f"  Permanent error calling Gemini API: {e}")
                    break

        if not success or not response:
            print(f"Failed to process row {idx + 1} after {max_retries} attempts.")
            continue

        try:
            parsed_data = None
            if hasattr(response, 'parsed') and response.parsed:
                parsed_data = response.parsed
            
            if parsed_data:
                df.at[idx, 'Work City'] = parsed_data.work_city or ""
                df.at[idx, 'Work Country'] = parsed_data.work_country or ""
                df.at[idx, 'Updated City'] = parsed_data.updated_city or ""
                df.at[idx, 'Updated Country'] = parsed_data.updated_country or ""
            else:
                import json
                data = json.loads(response.text)
                df.at[idx, 'Work City'] = data.get('work_city') or ""
                df.at[idx, 'Work Country'] = data.get('work_country') or ""
                df.at[idx, 'Updated City'] = data.get('updated_city') or ""
                df.at[idx, 'Updated Country'] = data.get('updated_country') or ""
            
        except Exception as e:
            print(f"Error calling Gemini API for row {idx + 1}: {e}")

        # Introduce short delay between requests to respect API rate limits (free tier)
        time.sleep(2)

    # Write df to excel temporarily to get basic spreadsheet
    df.to_excel(output_file, index=False)

    # Use openpyxl to apply formatting and highlighting
    print("Applying highlighting to mismatches...")
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Find the column index for each column name
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    work_city_idx = header_map.get('Work City')
    work_country_idx = header_map.get('Work Country')
    updated_city_idx = header_map.get('Updated City')
    updated_country_idx = header_map.get('Updated Country')

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for r in range(2, ws.max_row + 1):
        # Read the cell values
        w_city = str(ws.cell(row=r, column=work_city_idx).value or "").strip().lower()
        u_city = str(ws.cell(row=r, column=updated_city_idx).value or "").strip().lower()
        w_country = str(ws.cell(row=r, column=work_country_idx).value or "").strip().lower()
        u_country = str(ws.cell(row=r, column=updated_country_idx).value or "").strip().lower()

        # Check mismatches
        city_mismatch = w_city != u_city
        country_mismatch = w_country != u_country

        if city_mismatch:
            ws.cell(row=r, column=work_city_idx).fill = yellow_fill
            ws.cell(row=r, column=updated_city_idx).fill = yellow_fill
            print(f"Mismatch highlighted for City in row {r}: '{ws.cell(row=r, column=work_city_idx).value}' vs '{ws.cell(row=r, column=updated_city_idx).value}'")

        if country_mismatch:
            ws.cell(row=r, column=work_country_idx).fill = yellow_fill
            ws.cell(row=r, column=updated_country_idx).fill = yellow_fill
            print(f"Mismatch highlighted for Country in row {r}: '{ws.cell(row=r, column=work_country_idx).value}' vs '{ws.cell(row=r, column=updated_country_idx).value}'")

    wb.save(output_file)
    print(f"Successfully saved highlighted mismatches to: {output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Standardize and match location columns in an Excel sheet.")
    parser.add_argument("--input", default="input.xlsx", help="Input Excel file path")
    parser.add_argument("--output", default="output.xlsx", help="Output Excel file path")
    args = parser.parse_args()

    standardize_and_match(args.input, args.output)
