import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Mock dependencies and DB values
class AlumniMock:
    company = "Supreme Committee for Delivery & Legacy"
    designation = "Policies & Standards Expert"
    location = "Qatar"
    education = [
        {'school_name': 'SP Jain School of Global Management - Dubai, Mumbai, Singapore & Sydney', 'degree': 'PGDM', 'field_of_study': 'Information Technology'}
    ]
    position_groups = None

alumni = AlumniMock()

file_path = "uploads/85_testt.xlsx"
df = pd.read_excel(file_path)

# Let's set some dummy values in df to test different cases
# row 0 has Name='Aditya Upreti', Work location (City & Country)='Doha, Qatar'
df.loc[0, 'Work location (City & Country)'] = 'Doha, Qatar'

orig_cols = [
    c
    for c in df.columns.tolist()
    if c.lower().replace(" ", "").replace("_", "")
    not in ["experience", "organisations"]
]

new_cols = [
    "New Company",
    "New Designation",
    "New Higher Education",
    "Updated Location",
    "Changed",
    "Position Groups",
]

actual_new_cols = [c for c in new_cols if c not in orig_cols]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Alumni Updates"

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

headers = orig_cols + actual_new_cols
ws.append(headers)

# Write header yellow fill
for col_idx in range(1, len(orig_cols) + 1):
    ws.cell(row=1, column=col_idx).fill = yellow_fill

column_map = {}
def normalize_header(name):
    if not name:
        return ""
    return str(name).strip().lower().replace(" ", "").replace("_", "").replace("-", "").replace("/", "").replace("(", "").replace(")", "").replace("&", "")

for col in df.columns:
    norm = normalize_header(col)
    if norm in ["companyorganizationname", "company", "currentcompany"]:
        column_map["company"] = col
    elif norm in ["designation", "currentdesignation"]:
        column_map["designation"] = col
    elif norm in ["worklocationcitycountry", "location", "worklocation"]:
        column_map["location"] = col
    elif norm in ["anyhigherqualificationreceivedafterpassingout", "highereducation", "higheredu", "qualification"]:
        column_map["higher_education"] = col
    elif norm in ["linkedinurl", "linkedinprofileurl"]:
        column_map["linkedin_url"] = col

def extract_country(location_str):
    if not location_str: return ""
    parts = [p.strip() for p in str(location_str).split(",") if p.strip()]
    return parts[-1] if parts else str(location_str)

def normalize_string(val):
    if not val or pd.isna(val): return ""
    import re
    val = str(val).strip().lower()
    if val in ("nan", "none", "null", ""): return ""
    val = re.sub(r"[^\w\s]", " ", val)
    val = re.sub(r"\s+", " ", val)
    return val.strip()

for _, row in df.iterrows():
    row_values = []
    for col in orig_cols:
        val = row[col]
        if pd.isna(val):
            val = ""
        row_values.append(val)
        
    def clean_nan(v):
        if not v or pd.isna(v) or str(v).lower().strip() == "nan":
            return ""
        return str(v).strip()

    old_company = clean_nan(row.get(column_map.get("company", ""), ""))
    old_designation = clean_nan(row.get(column_map.get("designation", ""), ""))
    old_location = clean_nan(row.get(column_map.get("location", ""), ""))
    old_higher_edu = clean_nan(row.get(column_map.get("higher_education", ""), ""))

    db_company = alumni.company or ""
    db_designation = alumni.designation or ""
    db_location = alumni.location or ""
    latest_higher_edu = "SP Jain School of Global Management - Dubai, Mumbai, Singapore & Sydney (PGDM) - Information Technology"
    position_groups_val = ""

    old_comp_norm = normalize_string(old_company)
    db_comp_norm = normalize_string(db_company)
    old_des_norm = normalize_string(old_designation)
    db_des_norm = normalize_string(db_designation)
    old_loc_norm = normalize_string(extract_country(old_location))
    db_loc_norm = normalize_string(extract_country(db_location))
    old_edu_norm = normalize_string(old_higher_edu)
    db_edu_norm = normalize_string(latest_higher_edu)

    company_changed = old_comp_norm != db_comp_norm
    designation_changed = old_des_norm != db_des_norm
    location_changed = old_loc_norm != db_loc_norm
    higher_edu_changed = old_edu_norm != db_edu_norm

    has_any_changed = company_changed or designation_changed or location_changed or higher_edu_changed

    new_vals_dict = {
        "New Company": db_company,
        "New Designation": db_designation,
        "New Higher Education": latest_higher_edu if higher_edu_changed else "",
        "Updated Location": db_location,
        "Changed": "True" if has_any_changed else "",
        "Position Groups": position_groups_val,
    }

    for col_name in actual_new_cols:
        row_values.append(new_vals_dict[col_name])

    ws.append(row_values)
    current_row = ws.max_row

    from openpyxl.styles import Alignment
    center_align = Alignment(horizontal="center", vertical="center")

    for col_name in ["New Company", "New Designation", "New Higher Education", "Updated Location", "Changed", "Position Groups"]:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            cell_val = new_vals_dict[col_name]
            ws.cell(row=current_row, column=col_idx).value = cell_val

            should_color = False
            if col_name == "New Company" and company_changed:
                should_color = True
            elif col_name == "New Designation" and designation_changed:
                should_color = True
            elif col_name == "New Higher Education" and higher_edu_changed:
                should_color = True
            elif col_name == "Updated Location" and location_changed:
                should_color = True
            elif col_name == "Changed" and cell_val == "True":
                should_color = True

            if should_color:
                ws.cell(row=current_row, column=col_idx).fill = yellow_fill

            if col_name == "Changed":
                ws.cell(row=current_row, column=col_idx).alignment = center_align

# Print cell values and colors for row 2
print("Row 2 cells:")
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=2, column=col_idx)
    color = cell.fill.start_color.rgb if cell.fill else None
    print(f"Col {col_idx} ({headers[col_idx-1]}): value='{cell.value}', color={color}")
