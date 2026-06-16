from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from typing import List
from app.core.database import get_db
from app.schemas.core_schemas import Alumni, AlumniCreate
from app.services import alumni_service
from app.services.apify_service import apify_service
from app.services.comparison_engine import comparison_engine
from fastapi import BackgroundTasks
from app.models import core_models
import logging

logger = logging.getLogger(__name__)

def track_alumni_pipeline(alumni_id: int, db: Session):
    try:
        logger.info(f"Starting tracking pipeline for alumni {alumni_id}")
        # Fetch alumni
        alumni = alumni_service.get_alumni(db, alumni_id)
        if not alumni or not alumni.linkedin_url:
            logger.error("Alumni or LinkedIn URL not found")
            return
            
        # Trigger Apify
        results = apify_service.enrich_linkedin_profile(alumni.linkedin_url)
        if not results:
            logger.error("Failed to fetch profile data from Apify")
            return
            
        # Compare and update
        comparison_engine.compare_and_update(db, alumni_id, results)
        logger.info(f"Tracking pipeline complete for alumni {alumni_id}")
        
    except Exception as e:
        logger.error(f"Error in track_alumni_pipeline: {str(e)}")
    finally:
        db.close()

router = APIRouter()

from typing import Optional
from fastapi.responses import StreamingResponse
import io
import pandas as pd

@router.get("/", response_model=List[Alumni])
def read_alumni_list(
    skip: int = 0, limit: int = 100, 
    upload_id: Optional[int] = None,
    q: Optional[str] = None, company: Optional[str] = None, location: Optional[str] = None,
    db: Session = Depends(get_db)
):
    alumni = alumni_service.get_alumni_list(db, skip=skip, limit=limit, query=q, company=company, location=location)
    
    # Filter by upload_id if provided
    if upload_id:
        upload_alumni_ids = db.query(core_models.UploadAlumni.alumni_id).filter(
            core_models.UploadAlumni.upload_id == upload_id
        ).all()
        upload_alumni_ids = [ua[0] for ua in upload_alumni_ids]
        alumni = [a for a in alumni if a.id in upload_alumni_ids]
    
    return alumni

@router.get("/export")
def export_alumni(
    upload_id: Optional[int] = None,
    q: Optional[str] = None, company: Optional[str] = None, location: Optional[str] = None,
    db: Session = Depends(get_db)
):
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Font
    import os
    import io
    from fastapi.responses import StreamingResponse
    import json
    import pandas as pd
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Get all matching without limit for export
    alumni = alumni_service.get_alumni_list(db, skip=0, limit=100000, query=q, company=company, location=location)
    
    # Filter by upload_id if provided
    if upload_id:
        upload_alumni_ids = db.query(core_models.UploadAlumni.alumni_id).filter(
            core_models.UploadAlumni.upload_id == upload_id
        ).all()
        upload_alumni_ids = [ua[0] for ua in upload_alumni_ids]
        alumni = [a for a in alumni if a.id in upload_alumni_ids]
        
    def normalize_string(val):
        if not val or pd.isna(val):
            return ""
        val_str = str(val).strip().lower()
        if val_str in ['nan', 'none', 'null', '']:
            return ""
        import re
        return re.sub(r'[^\w\s]', '', val_str)
        
    def get_latest_position_info(position_groups):
        if not position_groups:
            return "", "", ""
        if isinstance(position_groups, str):
            try:
                position_groups = json.loads(position_groups)
            except Exception:
                return "", "", ""
        if not isinstance(position_groups, list) or len(position_groups) == 0:
            return "", "", ""
        first_group = position_groups[0]
        if not isinstance(first_group, dict):
            return "", "", ""
        latest_company = first_group.get('companyName') or first_group.get('company') or ""
        latest_designation = ""
        latest_location = ""
        positions = first_group.get('profilePositions', [])
        if isinstance(positions, list) and len(positions) > 0:
            first_pos = positions[0]
            if isinstance(first_pos, dict):
                latest_designation = first_pos.get('title') or ""
                latest_location = first_pos.get('location') or ""
        if not latest_designation:
            latest_designation = first_group.get('title') or ""
        if not latest_location:
            latest_location = first_group.get('location') or ""
        return latest_company, latest_designation, latest_location

    def extract_country(location_str):
        if not location_str:
            return ""
        val_str = str(location_str).strip()
        if not val_str or val_str.lower() in ('nan', 'none', 'null', ''):
            return ""
        if val_str.startswith('{') or val_str.startswith('['):
            try:
                parsed = json.loads(val_str)
                if isinstance(parsed, list) and parsed:
                    parsed = parsed[0]
                if isinstance(parsed, dict):
                    for key in ['country', 'countryName', 'country_code', 'location', 'name', 'city']:
                        if key in parsed and parsed[key]:
                            return extract_country(parsed[key])
            except Exception:
                try:
                    import ast
                    parsed = ast.literal_eval(val_str)
                    if isinstance(parsed, list) and parsed:
                        parsed = parsed[0]
                    if isinstance(parsed, dict):
                        for key in ['country', 'countryName', 'country_code', 'location', 'name', 'city']:
                            if key in parsed and parsed[key]:
                                return extract_country(parsed[key])
                except Exception:
                    pass
        parts = [p.strip() for p in val_str.split(',') if p.strip()]
        if parts:
            return parts[-1]
        return val_str

    def get_full_location(location_str):
        if not location_str:
            return ""
        val_str = str(location_str).strip()
        if not val_str or val_str.lower() in ('nan', 'none', 'null', ''):
            return ""
        if val_str.startswith('{') or val_str.startswith('['):
            try:
                parsed = json.loads(val_str)
                if isinstance(parsed, list) and parsed:
                    parsed = parsed[0]
                if isinstance(parsed, dict):
                    if 'location' in parsed and parsed['location']:
                        return get_full_location(parsed['location'])
                    for key in ['full_name', 'name', 'city']:
                        if key in parsed and parsed[key]:
                            return get_full_location(parsed[key])
            except Exception:
                try:
                    import ast
                    parsed = ast.literal_eval(val_str)
                    if isinstance(parsed, list) and parsed:
                        parsed = parsed[0]
                    if isinstance(parsed, dict):
                        if 'location' in parsed and parsed['location']:
                            return get_full_location(parsed['location'])
                        for key in ['full_name', 'name', 'city']:
                            if key in parsed and parsed[key]:
                                return get_full_location(parsed[key])
                except Exception:
                    pass
        return val_str

    # Try template-based export if upload_id is provided
    if upload_id:
        uploaded_file = db.query(core_models.UploadedFile).filter(core_models.UploadedFile.id == upload_id).first()
        if uploaded_file:
            file_path = os.path.join("uploads", f"{uploaded_file.id}_{uploaded_file.filename}")
            if os.path.exists(file_path) and file_path.lower().endswith(('.xlsx', '.xls')):
                try:
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    
                    # 1. Normalize headers and find target columns
                    def normalize_header(name):
                        if not name:
                            return ""
                        return str(name).lower().replace(" ", "").replace("_", "").replace("-", "").replace("/", "").replace("(", "").replace(")", "").replace(":", "")
                        
                    col_map = {}
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(row=1, column=c).value
                        norm = normalize_header(val)
                        if norm == 'linkedinurl' or norm == 'linkedinprofileurl':
                            col_map['linkedin_url'] = c
                        elif norm == 'name' or norm == 'fullname':
                            col_map['name'] = c
                        elif norm in ('company', 'companyorganizationname', 'currentcompany'):
                            col_map['company'] = c
                        elif norm == 'designation' or norm == 'currentdesignation':
                            col_map['designation'] = c
                        elif norm in ('worklocationcitycountry', 'location', 'worklocation'):
                            col_map['location'] = c
                        elif norm in ('anyhigherqualificationreceivedafterpassingout', 'anyhigherqualificationreceivedafterpassingoutfromiiitapleaseprovidedetails', 'highereducation', 'higheredu', 'qualification'):
                            col_map['higher_qual'] = c
                        elif norm == 'newcompany':
                            col_map['new_company'] = c
                        elif norm == 'newdesignation':
                            col_map['new_designation'] = c
                        elif norm == 'updatedlocation' or norm == 'newlocation':
                            col_map['updated_location'] = c
                        elif norm == 'newhighereducation' or norm == 'newhigherqual':
                            col_map['new_higher_edu'] = c
                        elif norm in ('flag', 'changed', 'tick', 'status'):
                            col_map['flag'] = c
                            
                    # Find empty highlighted column N/O as flag if not found
                    if 'flag' not in col_map:
                        for c in range(1, ws.max_column + 1):
                            cell = ws.cell(row=1, column=c)
                            val = cell.value
                            fill = cell.fill
                            color = fill.start_color.rgb if fill and fill.fill_type and fill.fill_type != 'none' else None
                            if color and color in ('FFFFFF00', '00FFFF00', 'FFFF00') and (val is None or str(val).lower().startswith('unnamed')):
                                col_map['flag'] = c
                                cell.value = "Flag"
                                break
                                
                    # Add missing tracking/flag columns if not exist
                    for field, name in [
                        ('new_company', 'New Company'),
                        ('new_designation', 'New Designation'),
                        ('updated_location', 'Updated Location'),
                        ('new_higher_edu', 'New Higher Education'),
                        ('flag', 'Changed')
                    ]:
                        if field not in col_map:
                            ws.max_column
                            new_col = ws.max_column + 1
                            ws.cell(row=1, column=new_col).value = name
                            ws.cell(row=1, column=new_col).fill = yellow_fill
                            col_map[field] = new_col
                            
                    # 2. Iterate rows and write updates
                    link_col = col_map.get('linkedin_url')
                    if link_col:
                        for r in range(2, ws.max_row + 1):
                            url_val = ws.cell(row=r, column=link_col).value
                            if not url_val:
                                continue
                            url_str = str(url_val).strip()
                            
                            alumni_rec = db.query(core_models.AlumniMaster).filter(
                                core_models.AlumniMaster.linkedin_url == url_str
                            ).first()
                            
                            has_changes = False
                            
                            def clean_val(v):
                                if not v or pd.isna(v) or str(v).lower().strip() == 'nan':
                                    return ""
                                return str(v).strip()
                                
                            old_comp = clean_val(ws.cell(row=r, column=col_map['company']).value) if 'company' in col_map else ""
                            old_des = clean_val(ws.cell(row=r, column=col_map['designation']).value) if 'designation' in col_map else ""
                            old_loc = extract_country(clean_val(ws.cell(row=r, column=col_map['location']).value)) if 'location' in col_map else ""
                            old_edu = clean_val(ws.cell(row=r, column=col_map['higher_qual']).value) if 'higher_qual' in col_map else ""
                            
                            if alumni_rec:
                                db_company = alumni_rec.company or ""
                                db_designation = alumni_rec.designation or ""
                                db_location = alumni_rec.location or ""
                                
                                latest_higher_edu = ""
                                if alumni_rec.education:
                                    edu_list = alumni_rec.education
                                    if isinstance(edu_list, str):
                                        try:
                                            edu_list = json.loads(edu_list)
                                        except Exception:
                                            edu_list = []
                                    edu_entries = []
                                    if isinstance(edu_list, list):
                                        for edu in edu_list:
                                            if not isinstance(edu, dict):
                                                continue
                                            school = edu.get("school_name") or edu.get("schoolName") or edu.get("school") or ""
                                            degree = edu.get("degree") or edu.get("degreeName") or ""
                                            field = edu.get("field_of_study") or edu.get("fieldOfStudy") or ""
                                            if school and 'iiit' in school.lower():
                                                continue
                                            entry = school
                                            if degree:
                                                entry += f" ({degree})"
                                            if field:
                                                entry += f" - {field}"
                                            if entry.strip():
                                                edu_entries.append(entry)
                                    latest_higher_edu = "; ".join(edu_entries)
                                    
                                if db_company:
                                    ws.cell(row=r, column=col_map['new_company']).value = db_company
                                    if normalize_string(old_comp) != normalize_string(db_company):
                                        ws.cell(row=r, column=col_map['new_company']).fill = yellow_fill
                                        has_changes = True
                                    
                                if db_designation:
                                    ws.cell(row=r, column=col_map['new_designation']).value = db_designation
                                    if normalize_string(old_des) != normalize_string(db_designation):
                                        ws.cell(row=r, column=col_map['new_designation']).fill = yellow_fill
                                        has_changes = True
                                    
                                if db_location:
                                    ws.cell(row=r, column=col_map['updated_location']).value = db_location
                                    if normalize_string(old_loc) != normalize_string(extract_country(db_location)):
                                        ws.cell(row=r, column=col_map['updated_location']).fill = yellow_fill
                                        has_changes = True
                                    
                                if normalize_string(old_edu) != normalize_string(latest_higher_edu) and latest_higher_edu:
                                    ws.cell(row=r, column=col_map['new_higher_edu']).value = latest_higher_edu
                                    ws.cell(row=r, column=col_map['new_higher_edu']).fill = yellow_fill
                                    has_changes = True
                                    
                            # Scan to see if any cell in the row is highlighted in yellow (excluding the flag cell itself)
                            has_highlight = False
                            for c in range(1, ws.max_column + 1):
                                if c == col_map['flag']:
                                    continue
                                cell = ws.cell(row=r, column=c)
                                fill = cell.fill
                                if fill and fill.fill_type and fill.fill_type != 'none':
                                    color = fill.start_color.rgb if fill.start_color else None
                                    if color:
                                        color_str = str(color).upper()
                                        if color_str in ('FFFF00', '00FFFF00', 'FFFFFF00', 'FFFFFFFF00') or color_str.endswith('FFFF00'):
                                            has_highlight = True
                                            break
                                        
                            flag_cell = ws.cell(row=r, column=col_map['flag'])
                            if has_highlight:
                                flag_cell.value = "True"
                                flag_cell.alignment = center_alignment
                            else:
                                flag_cell.value = ""
                                
                    stream = io.BytesIO()
                    wb.save(stream)
                    stream.seek(0)
                    
                    response = StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    response.headers["Content-Disposition"] = f"attachment; filename={uploaded_file.filename}"
                    return response
                except Exception as ex:
                    logger.error(f"Template export failed, falling back to scratch export: {ex}")

    # Fallback to scratch export
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alumni Updates"
    
    # Pre-calculate max changes per field across all exported alumni
    max_changes = {'full_name': 0, 'company': 0, 'designation': 0, 'location': 0}
    alumni_logs = {}
    
    for a in alumni:
        logs = db.query(core_models.ChangeLog).filter(
            core_models.ChangeLog.alumni_id == a.id,
            core_models.ChangeLog.field_changed.in_(['full_name', 'company', 'designation', 'location'])
        ).order_by(core_models.ChangeLog.detected_at.asc()).all()
        alumni_logs[a.id] = logs
        counts = {'full_name': 0, 'company': 0, 'designation': 0, 'location': 0}
        for log in logs:
            counts[log.field_changed] += 1
            
        for k in counts:
            if counts[k] > max_changes[k]:
                max_changes[k] = counts[k]
                
    # Build dynamic headers
    field_to_header = {
        'full_name': 'Name',
        'company': 'Company/Organization Name',
        'designation': 'Designation',
        'location': 'Work location (City & Country)'
    }
    
    columns = [
        "Name", "Enrollment No.", "Course completed from IIITA (Highest Degree)", "Year of Passing",
        "linkedin_url", "Designation", "Company/Organization Name", "Work location (City & Country)",
        "Any Higher Qualification received after passing out"
    ]
    
    dynamic_fields_order = ['full_name', 'company', 'designation', 'location']
    for field in dynamic_fields_order:
        for _ in range(max_changes[field]):
            columns.append(field_to_header[field])
            
    columns.extend(["Changed", "Publications"])
    ws.append(columns)
    
    for a in alumni:
        extra = a.extra_data if a.extra_data else {}
        enrollment_no = extra.get('Enrollment No.', '')
        
        course = extra.get('Course completed from IIITA (Highest Degree)', '')
        year = extra.get('Year of Passing', '')
        higher_qual = extra.get('Higher Qualification', '')
        
        if not higher_qual and a.education:
            try:
                education_list = a.education if isinstance(a.education, list) else [a.education]
                if isinstance(education_list, str):
                    education_list = json.loads(education_list)
                    if not isinstance(education_list, list):
                        education_list = [education_list]
                
                qualifications = []
                for edu in education_list:
                    if isinstance(edu, dict):
                        school = edu.get('school_name') or edu.get('schoolName') or 'Unknown Institution'
                        degree = edu.get('degree') or edu.get('fieldOfStudy') or 'Degree'
                        qualifications.append(f"{degree} from {school}")
                if qualifications:
                    higher_qual = ' | '.join(qualifications)
            except:
                pass
        
        mismatch_flag = ""
        if course and higher_qual:
            degree_part = higher_qual.split(' from ')[0].strip() if ' from ' in higher_qual else higher_qual
            course_normalized = course.lower().strip()
            degree_normalized = degree_part.lower().strip()
            if course_normalized != degree_normalized:
                course_words = set(course_normalized.split())
                degree_words = set(degree_normalized.split())
                common_words = course_words & degree_words
                meaningful_words = course_words - {'of', 'in', 'and', 'the', 'from', 'a', 'an'}
                if meaningful_words and len(common_words) < len(meaningful_words) * 0.5:
                    mismatch_flag = "⚠️ MISMATCH"
        
        f_logs = [l for l in alumni_logs[a.id] if l.field_changed == 'full_name']
        old_name = extra.get('Name') or (f_logs[0].old_value if f_logs else a.full_name) or ''
        
        c_logs = [l for l in alumni_logs[a.id] if l.field_changed == 'company']
        old_company = extra.get('Company/Organization Name') or (c_logs[0].old_value if c_logs else a.company) or ''
        
        d_logs = [l for l in alumni_logs[a.id] if l.field_changed == 'designation']
        old_designation = extra.get('Designation') or (d_logs[0].old_value if d_logs else a.designation) or ''
        
        l_logs = [l for l in alumni_logs[a.id] if l.field_changed == 'location']
        old_location = get_full_location(extra.get('Work location (City & Country)') or (l_logs[0].old_value if l_logs else a.location) or '')
        
        row_data = [
            old_name, enrollment_no, course, year, a.linkedin_url,
            old_designation, old_company, old_location, higher_qual
        ]
        
        dynamic_colors = []
        for field in dynamic_fields_order:
            logs_for_field = [l for l in alumni_logs[a.id] if l.field_changed == field]
            for i in range(max_changes[field]):
                if i < len(logs_for_field):
                    val = logs_for_field[i].new_value
                    if field == 'location':
                        val = get_full_location(val)
                    row_data.append(val)
                    dynamic_colors.append(True)
                else:
                    row_data.append("")
                    dynamic_colors.append(False)
                    
        def format_field(data, field_type):
            if not data:
                return ""
            if isinstance(data, str):
                try:
                    data = json.loads(data)
                except Exception:
                    return data
            if not isinstance(data, list):
                return str(data)
                
            lines = []
            for item in data:
                if not isinstance(item, dict):
                    continue
                def get_str(val):
                    if isinstance(val, dict):
                        return val.get('name', '')
                    return str(val) if val else ''
                def get_date(val):
                    if isinstance(val, dict):
                        start = val.get('start', {}).get('year') if isinstance(val.get('start'), dict) else val.get('start')
                        end = val.get('end', {}).get('year') if isinstance(val.get('end'), dict) else val.get('end')
                        if start and end:
                            return f"{start} - {end}"
                        return str(start or end or '')
                    return str(val) if val else ''
                if field_type == 'position_groups':
                    company = get_str(item.get('companyName') or item.get('company') or 'Unknown Company')
                    start_date = item.get('start_date') or item.get('startDate') or item.get('start') or ''
                    end_date = item.get('end_date') or item.get('endDate') or item.get('end') or ''
                    if start_date and end_date:
                        lines.append(f"{company} ({start_date} - {end_date})")
                    elif start_date or end_date:
                        lines.append(f"{company} ({start_date or end_date})")
                    else:
                        lines.append(company)
                else:
                    title = get_str(item.get('title') or item.get('name') or item.get('companyName') or 'Untitled')
                    company = get_str(item.get('company') or item.get('publisher') or item.get('position') or '')
                    date_str = get_date(item.get('dates') or item.get('dateRange') or item.get('date') or item.get('year'))
                    if not date_str:
                        start_date = item.get('start_date') or item.get('startDate') or item.get('start') or ''
                        end_date = item.get('end_date') or item.get('endDate') or item.get('end') or ''
                        if start_date and end_date:
                            date_str = f"{start_date} - {end_date}"
                        elif start_date or end_date:
                            date_str = start_date or end_date
                    location = extract_country(get_str(item.get('location')))
                    parts = []
                    if title: parts.append(title)
                    if company and company != title: parts.append(f"at {company}")
                    if date_str: parts.append(f"({date_str})")
                    if location: parts.append(f"- {location}")
                    lines.append(" ".join(parts))
                positions = item.get('profilePositions', [])
                if isinstance(positions, list) and positions:
                    for pos in positions:
                        if not isinstance(pos, dict):
                            continue
                        pos_title = get_str(pos.get('title'))
                        pos_date = get_date(pos.get('dateRange') or pos.get('dates'))
                        if not pos_date:
                            start_date = pos.get('start_date') or pos.get('startDate') or pos.get('start') or ''
                            end_date = pos.get('end_date') or pos.get('endDate') or pos.get('end') or ''
                            if start_date and end_date:
                                pos_date = f"{start_date} - {end_date}"
                            elif start_date or end_date:
                                pos_date = start_date or end_date
                        pos_loc = extract_country(get_str(pos.get('location')))
                        p_parts = [f"  -> {pos_title}"]
                        if pos_date: p_parts.append(f"({pos_date})")
                        if pos_loc: p_parts.append(f"- {pos_loc}")
                        lines.append(" ".join(p_parts))
            return "\n".join(lines)
            
        has_any_highlight = any(dynamic_colors) or bool(mismatch_flag)
        row_data.extend([
            "True" if has_any_highlight else "",
            format_field(a.publications, 'publications')
        ])
        
        ws.append(row_data)
        current_row = ws.max_row
        
        base_col_len = 9
        for i, color_it in enumerate(dynamic_colors):
            if color_it:
                ws.cell(row=current_row, column=base_col_len + 1 + i).fill = yellow_fill
                
        if mismatch_flag:
            qual_cell = ws.cell(row=current_row, column=9)
            qual_cell.fill = yellow_fill
            qual_cell.font = Font(bold=True, color="000000")
            
        # Center align Changed tick
        changed_col_idx = base_col_len + len(dynamic_colors) + 1
        ws.cell(row=current_row, column=changed_col_idx).alignment = center_alignment
        
    try:
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        response = StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers["Content-Disposition"] = "attachment; filename=Alumni_Updates.xlsx"
        return response
    except Exception as e:
        logger.error(f"Error creating Excel export: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to create export file: {str(e)}"
        )

from fastapi import Request
from app.core.limiter import limiter

@router.post("/", response_model=Alumni)
def create_alumni(alumni: AlumniCreate, db: Session = Depends(get_db)):
    db_alumni = alumni_service.get_alumni_by_url(db, linkedin_url=alumni.linkedin_url)
    if db_alumni:
        raise HTTPException(status_code=400, detail="Alumni with this LinkedIn URL already exists")
    return alumni_service.create_alumni(db=db, alumni=alumni)

@router.post("/{alumni_id}/track")
def track_alumni(alumni_id: int, db: Session = Depends(get_db)):
    alumni = alumni_service.get_alumni(db, alumni_id=alumni_id)
    if not alumni:
        raise HTTPException(status_code=404, detail="Alumni not found")
        
    # Run synchronously with a dedicated session (since pipeline closes it)
    from app.core.database import SessionLocal
    sync_db = SessionLocal()
    track_alumni_pipeline(alumni_id, sync_db)
    
    return {"message": "Tracking pipeline completed successfully", "alumni_id": alumni_id}

@router.post("/add-single", response_model=Alumni)
@limiter.limit("5/minute")
def add_single_alumni(request: Request, alumni: AlumniCreate, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    db_alumni = alumni_service.get_alumni_by_url(db, linkedin_url=alumni.linkedin_url)
    if db_alumni:
        raise HTTPException(status_code=400, detail="Alumni with this LinkedIn URL already exists")
    
    created = alumni_service.create_alumni(db=db, alumni=alumni)
    
    from app.core.database import SessionLocal
    background_tasks.add_task(track_alumni_pipeline, created.id, SessionLocal())
    
    return created

@router.get("/search", response_model=List[Alumni])
@limiter.limit("30/minute")
def search_alumni(request: Request, q: str, company: Optional[str] = None, location: Optional[str] = None, db: Session = Depends(get_db)):
    return alumni_service.get_alumni_list(db, query=q, company=company, location=location)

@router.get("/{alumni_id}", response_model=Alumni)
def read_alumni(alumni_id: int, db: Session = Depends(get_db)):
    db_alumni = alumni_service.get_alumni(db, alumni_id=alumni_id)
    if db_alumni is None:
        raise HTTPException(status_code=404, detail="Alumni not found")
    return db_alumni

from app.models import core_models
@router.delete("/{alumni_id}")
def delete_alumni(alumni_id: int, db: Session = Depends(get_db)):
    alumni = db.query(core_models.AlumniMaster).filter(core_models.AlumniMaster.id == alumni_id).first()
    if not alumni:
        raise HTTPException(status_code=404, detail="Alumni not found")
        
    # Delete related tracking logs first
    db.query(core_models.ChangeLog).filter(core_models.ChangeLog.alumni_id == alumni_id).delete()
    db.query(core_models.AlumniHistory).filter(core_models.AlumniHistory.alumni_id == alumni_id).delete()
    
    db.delete(alumni)
    db.commit()
    return {"message": "Alumni deleted successfully"}

@router.get("/{alumni_id}/changelog")
def get_alumni_changelog(alumni_id: int, db: Session = Depends(get_db)):
    results = db.query(core_models.ChangeLog).filter(core_models.ChangeLog.alumni_id == alumni_id).order_by(core_models.ChangeLog.detected_at.desc()).all()
    return results
