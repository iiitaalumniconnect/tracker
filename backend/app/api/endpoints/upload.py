from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
import pandas as pd
import io
import os
from typing import List
from app.core.database import get_db, SessionLocal
from app.services import alumni_service
from app.schemas.core_schemas import AlumniCreate, UploadedFileSchema
from app.models.core_models import UploadedFile, UploadAlumni
from app.api.endpoints import alumni

router = APIRouter()

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


def process_csv_background(file_path: str, record_id: int):
    db = SessionLocal()

    try:
        if file_path.lower().endswith((".xlsx", ".xls")):
            df = pd.read_excel(file_path)
        else:
            df = pd.read_csv(file_path)

        imported_count = 0

        for _, row in df.iterrows():

            linkedin_url = str(row.get("linkedin_url", "")).strip()

            if not linkedin_url or linkedin_url.lower() == "nan":
                continue

            existing = alumni_service.get_alumni_by_url(db, linkedin_url)

            row_dict = row.to_dict()

            full_name = str(row_dict.pop("Name", row_dict.pop("full_name", "")))

            company = str(
                row_dict.pop("Company/Organization Name", row_dict.pop("company", ""))
            )

            designation = str(
                row_dict.pop("Designation", row_dict.pop("designation", ""))
            )

            location = str(
                row_dict.pop(
                    "Work location (City & Country)", row_dict.pop("location", "")
                )
            )

            location = extract_country(location)

            highest_degree = str(
                row_dict.get("Course completed from IIITA (Highest Degree)", "")
            ).strip()

            if "linkedin_url" in row_dict:
                del row_dict["linkedin_url"]

            import math

            extra_data = {
                k: (v if not (isinstance(v, float) and math.isnan(v)) else "")
                for k, v in row_dict.items()
            }

            if not existing:
                # Start with empty education - will be populated by API
                education = []

                alumni_create = AlumniCreate(
                    linkedin_url=linkedin_url,
                    full_name=full_name if full_name.lower() != "nan" else "",
                    company=company if company.lower() != "nan" else "",
                    designation=designation if designation.lower() != "nan" else "",
                    location=location if location.lower() != "nan" else "",
                    education=education,
                    extra_data=extra_data,
                )

                created_alumni = alumni_service.create_alumni(db, alumni_create)

                # Track the alumni-upload mapping
                upload_alumni_record = UploadAlumni(
                    upload_id=record_id, alumni_id=created_alumni.id
                )
                db.add(upload_alumni_record)
                db.commit()

                imported_count += 1
            else:
                # Update existing alumni record - add to extra_data only
                updated = False

                # Update missing or empty fields in extra_data
                existing_extra = existing.extra_data
                if isinstance(existing_extra, str):
                    import json

                    try:
                        existing_extra = json.loads(existing_extra)
                    except Exception:
                        existing_extra = {}
                if not isinstance(existing_extra, dict):
                    existing_extra = {}

                for k, v in extra_data.items():
                    if k not in existing_extra or not existing_extra[k]:
                        existing_extra[k] = v
                        updated = True

                if updated:
                    from sqlalchemy.orm.attributes import flag_modified

                    existing.extra_data = existing_extra
                    flag_modified(existing, "extra_data")
                    db.add(existing)
                    db.commit()

                # Always track the alumni-upload mapping (even for existing records)
                existing_mapping = (
                    db.query(UploadAlumni)
                    .filter(
                        UploadAlumni.upload_id == record_id,
                        UploadAlumni.alumni_id == existing.id,
                    )
                    .first()
                )

                if not existing_mapping:
                    upload_alumni_record = UploadAlumni(
                        upload_id=record_id, alumni_id=existing.id
                    )
                    db.add(upload_alumni_record)
                    db.commit()

                imported_count += 1

        record = db.query(UploadedFile).filter(UploadedFile.id == record_id).first()

        if record:
            record.status = "completed"
            record.record_count = imported_count
            db.commit()

            import asyncio
            from app.core.ws_manager import manager

            try:
                asyncio.run(
                    manager.broadcast(
                        {
                            "type": "status_update",
                            "file_id": record_id,
                            "status": "completed",
                        }
                    )
                )
            except Exception:
                pass

    except Exception as e:

        record = db.query(UploadedFile).filter(UploadedFile.id == record_id).first()

        if record:
            record.status = f"failed: {str(e)}"
            db.commit()

            import asyncio
            from app.core.ws_manager import manager

            try:
                asyncio.run(
                    manager.broadcast(
                        {
                            "type": "status_update",
                            "file_id": record_id,
                            "status": "failed",
                        }
                    )
                )
            except Exception:
                pass

    finally:
        db.close()


@router.post("/upload-csv")
async def upload_csv(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be a CSV or Excel file")

    contents = await file.read()
    try:
        if file.filename.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(contents))
        else:
            df = pd.read_csv(io.StringIO(contents.decode("utf-8")))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")

    if len(df) > 1000:
        raise HTTPException(
            status_code=400,
            detail=f"File exceeds the maximum limit of 1000 rows. You provided {len(df)} rows.",
        )

    required_cols = ["linkedin_url"]
    for col in required_cols:
        if col not in df.columns:
            raise HTTPException(
                status_code=400, detail=f"Missing required column: {col}"
            )

    # Create uploaded file record
    uploaded_record = UploadedFile(
        filename=file.filename,
        status="processing",
        record_count=0,  # Will update when finished
    )
    db.add(uploaded_record)
    db.commit()
    db.refresh(uploaded_record)

    # Save file to disk
    file_path = os.path.join(UPLOAD_DIR, f"{uploaded_record.id}_{file.filename}")
    with open(file_path, "wb") as f:
        f.write(contents)

    # Trigger background task
    background_tasks.add_task(process_csv_background, file_path, uploaded_record.id)

    return {
        "message": f"Upload accepted. Processing {len(df)} records in the background.",
        "file_id": uploaded_record.id,
    }


@router.get("/history", response_model=List[UploadedFileSchema])
def get_upload_history(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    files = (
        db.query(UploadedFile)
        .order_by(UploadedFile.uploaded_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )
    return files


@router.delete("/history/{file_id}")
def delete_upload_history(file_id: int, db: Session = Depends(get_db)):
    file_record = db.query(UploadedFile).filter(UploadedFile.id == file_id).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="File history not found")

    try:
        # First, delete all UploadAlumni records associated with this file
        # This prevents foreign key constraint violation
        db.query(UploadAlumni).filter(UploadAlumni.upload_id == file_id).delete()

        # Then delete the UploadedFile record
        db.delete(file_record)
        db.commit()

        # Finally, delete the physical file if it exists
        file_path = os.path.join(UPLOAD_DIR, f"{file_record.id}_{file_record.filename}")
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as e:
                print(f"Warning: Could not delete physical file {file_path}: {str(e)}")

        return {"message": "File history and associated records deleted successfully"}

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500, detail=f"Failed to delete history record: {str(e)}"
        )


import json


def extract_country(location_str):
    if not location_str:
        return ""

    val_str = str(location_str).strip()

    if not val_str or val_str.lower() in ("nan", "none", "null", ""):
        return ""

    # Check if it looks like a JSON object or list
    if val_str.startswith("{") or val_str.startswith("["):
        try:
            import json

            parsed = json.loads(val_str)
            if isinstance(parsed, list) and parsed:
                parsed = parsed[0]
            if isinstance(parsed, dict):
                # Search for country keys in preferred order
                for key in [
                    "country",
                    "countryName",
                    "country_code",
                    "location",
                    "name",
                    "city",
                ]:
                    if key in parsed and parsed[key]:
                        # Recursively extract country from the found value
                        return extract_country(parsed[key])
        except Exception:
            # If standard json fails, try ast.literal_eval
            try:
                import ast

                parsed = ast.literal_eval(val_str)
                if isinstance(parsed, list) and parsed:
                    parsed = parsed[0]
                if isinstance(parsed, dict):
                    for key in [
                        "country",
                        "countryName",
                        "country_code",
                        "location",
                        "name",
                        "city",
                    ]:
                        if key in parsed and parsed[key]:
                            return extract_country(parsed[key])
            except Exception:
                pass

    parts = [p.strip() for p in val_str.split(",") if p.strip()]
    return parts[-1] if parts else val_str


def normalize_string(val):
    if not val or pd.isna(val):
        return ""
    val_str = str(val).strip().lower()
    if val_str in ["nan", "none", "null", ""]:
        return ""
    import re

    return re.sub(r"[^\w\s]", "", val_str)


def get_latest_position_info(position_groups):
    if not position_groups:
        return "", "", ""

    if isinstance(position_groups, str):
        try:
            position_groups = json.loads(position_groups)
        except Exception:
            return "", "", ""

    if not isinstance(position_groups, list) or len(position_groups) == 0:
        return "", "", ""

    first_group = position_groups[0]
    if not isinstance(first_group, dict):
        return "", "", ""

    latest_company = first_group.get("companyName") or first_group.get("company") or ""
    latest_designation = ""
    latest_location = ""

    positions = first_group.get("profilePositions", [])
    if isinstance(positions, list) and len(positions) > 0:
        first_pos = positions[0]
        if isinstance(first_pos, dict):
            latest_designation = first_pos.get("title") or ""
            latest_location = first_pos.get("location") or ""

    if not latest_designation:
        latest_designation = first_group.get("title") or ""
    if not latest_location:
        latest_location = first_group.get("location") or ""

    return latest_company, latest_designation, latest_location


def format_field(data, field_type):
    if not data:
        return ""

    if isinstance(data, str):
        try:
            data = json.loads(data)
        except Exception:
            return data

    if not isinstance(data, list):
        return str(data)

    def get_str(val):
        if isinstance(val, dict):
            return val.get("name", "")
        return str(val) if val else ""

    # Full formatting for position_groups listing companies and sub-positions (titles only)
    if field_type == "position_groups":
        lines = []
        for item in data:
            if not isinstance(item, dict):
                continue
            company = get_str(
                item.get("companyName") or item.get("company") or "Unknown Company"
            )

            # Try to extract start_date and end_date
            start = (
                item.get("start_date")
                or item.get("startDate")
                or item.get("start")
                or ""
            )
            end = item.get("end_date") or item.get("endDate") or item.get("end") or ""
            date_str = f" ({start} - {end})" if (start or end) else ""

            lines.append(f"{company}{date_str}")

            positions = item.get("profilePositions", [])
            if isinstance(positions, list) and positions:
                for pos in positions:
                    if not isinstance(pos, dict):
                        continue
                    pos_title = get_str(pos.get("title"))
                    pos_start = (
                        pos.get("startDate")
                        or pos.get("start")
                        or pos.get("start_date")
                        or ""
                    )
                    pos_end = (
                        pos.get("endDate")
                        or pos.get("end")
                        or pos.get("end_date")
                        or ""
                    )
                    pos_date = (
                        f" ({pos_start} - {pos_end})" if (pos_start or pos_end) else ""
                    )
                    if pos_title:
                        lines.append(f"  -> {pos_title}{pos_date}")
        return "\n".join(lines)

    # Publications formatting
    elif field_type == "publications":
        lines = []
        for item in data:
            if not isinstance(item, dict):
                continue

            title = get_str(item.get("title") or item.get("name"))
            if title:
                lines.append(title)

        return "\n".join(lines)

    return ""


@router.get("/download/{file_id}")
def download_upload_history(file_id: int, db: Session = Depends(get_db)):
    file_record = db.query(UploadedFile).filter(UploadedFile.id == file_id).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="File history not found")

    file_path = os.path.join(UPLOAD_DIR, f"{file_record.id}_{file_record.filename}")
    if not os.path.exists(file_path):
        raise HTTPException(
            status_code=404, detail="Physical file no longer exists on the server"
        )

    import openpyxl
    from openpyxl.styles import PatternFill
    from app.models import core_models
    import io
    from fastapi.responses import StreamingResponse

    try:
        if file_path.lower().endswith((".xlsx", ".xls")):
            df = pd.read_excel(file_path)
        else:
            df = pd.read_csv(file_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read file: {str(e)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alumni Updates"

    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )

    # Remove unwanted columns - only exclude 'experience' and 'organisations'
    orig_cols = [
        c
        for c in df.columns.tolist()
        if c.lower().replace(" ", "").replace("_", "")
        not in ["experience", "organisations"]
    ]

    new_cols = [
        "New Company",
        "New Designation",
        "New Higher Education",
        "Updated Location",
        "Changed",
        "Position Groups",
    ]

    actual_new_cols = [c for c in new_cols if c not in orig_cols]

    headers = orig_cols + actual_new_cols
    ws.append(headers)

    for col_idx in range(1, len(orig_cols) + 1):
        ws.cell(row=1, column=col_idx).fill = yellow_fill

    for _, row in df.iterrows():
        linkedin_url = str(row.get("linkedin_url", "")).strip()

        alumni = None
        if linkedin_url and linkedin_url.lower() != "nan":
            alumni = (
                db.query(core_models.AlumniMaster)
                .filter(core_models.AlumniMaster.linkedin_url == linkedin_url)
                .first()
            )

        row_values = []

        for col in orig_cols:
            val = row[col]

            if pd.isna(val):
                val = ""

            col_clean = (
                col.lower()
                .replace("_", "")
                .replace(" ", "")
                .replace("(", "")
                .replace(")", "")
            )

            if col_clean in ["worklocation(city&country)", "location"]:
                val = extract_country(val)

            # Fill IIITA degree column from database
            if alumni and (
                "coursecompletedfromiiita" in col_clean
                or "coursecompletedfromiita" in col_clean
                or ("coursecompleted" in col_clean and "highestdegree" in col_clean)
            ):
                if alumni.education:
                    edu_list = alumni.education

                    if isinstance(edu_list, str):
                        try:
                            edu_list = json.loads(edu_list)
                        except Exception:
                            edu_list = []

                    if isinstance(edu_list, list):
                        for edu in edu_list:
                            if not isinstance(edu, dict):
                                continue

                            school_name = (
                                edu.get("school_name")
                                or edu.get("schoolName")
                                or edu.get("school")
                                or ""
                            )

                            # Look for IIITA entry
                            if school_name and "iiit" in school_name.lower():
                                degree = (
                                    edu.get("degree") or edu.get("degreeName") or ""
                                )
                                if degree:
                                    val = degree
                                break

            # Fill higher education column from database
            elif alumni and (
                "anyhigherqualification" in col_clean
                or "highereducation" in col_clean
                or "higheredu" in col_clean
                or "qualification" in col_clean
                or col_clean
                in [
                    "anyhigherqualificationreceivedafterpassingout",
                    "highereducation",
                    "higheredu",
                    "qualification",
                ]
            ):
                if alumni.education:
                    edu_list = alumni.education

                    if isinstance(edu_list, str):
                        try:
                            edu_list = json.loads(edu_list)
                        except Exception:
                            edu_list = []

                    if isinstance(edu_list, list):
                        edu_entries = []
                        for edu in edu_list:
                            if not isinstance(edu, dict):
                                continue

                            school_name = (
                                edu.get("school_name")
                                or edu.get("schoolName")
                                or edu.get("school")
                                or ""
                            )
                            degree = edu.get("degree") or edu.get("degreeName") or ""

                            # Exclude IIIT entries from higher education
                            if school_name and "iiit" in school_name.lower():
                                continue

                            if school_name:
                                entry = school_name
                                if degree:
                                    entry += f" ({degree})"
                                edu_entries.append(entry)

                        if edu_entries:
                            val = "; ".join(edu_entries)

            if alumni and col_clean in ["publications", "positiongroups"]:
                db_field = (
                    "position_groups" if col_clean == "positiongroups" else col_clean
                )
                db_val = getattr(alumni, db_field, None)

                if db_val:
                    val = format_field(db_val, db_field)

            row_values.append(val)

        # Safely extract old values (uploaded data)
        def clean_nan(v):
            if not v or pd.isna(v) or str(v).lower().strip() == "nan":
                return ""
            return str(v).strip()

        old_company = clean_nan(
            row.get("Company/Organization Name", row.get("company", ""))
        )
        old_designation = clean_nan(row.get("Designation", row.get("designation", "")))
        old_location = extract_country(
            clean_nan(
                row.get("Work location (City & Country)", row.get("location", ""))
            )
        )

        # Look up higher education column in original file
        higher_edu_col = next(
            (
                c
                for c in df.columns
                if "anyhigherqualification"
                in c.lower().replace(" ", "").replace("_", "")
                or "highereducation" in c.lower().replace(" ", "").replace("_", "")
                or "higheredu" in c.lower().replace(" ", "").replace("_", "")
                or c.lower().replace(" ", "").replace("_", "")
                in [
                    "anyhigherqualificationreceivedafterpassingout",
                    "highereducation",
                    "higheredu",
                    "qualification",
                ]
            ),
            None,
        )
        old_higher_edu = (
            clean_nan(row.get(higher_edu_col, "")) if higher_edu_col else ""
        )

        db_company = ""
        db_designation = ""
        db_location = ""
        latest_higher_edu = ""
        position_groups_val = ""

        company_changed = False
        designation_changed = False
        location_changed = False
        higher_edu_changed = False

        if alumni:
            latest_company, latest_designation, latest_location = (
                get_latest_position_info(alumni.position_groups)
            )

            db_company = latest_company or alumni.company or ""
            db_designation = latest_designation or alumni.designation or ""
            db_location = extract_country(latest_location or alumni.location or "")

            latest_higher_edu = ""

            if alumni.education:

                edu_list = alumni.education

                if isinstance(edu_list, str):
                    try:
                        edu_list = json.loads(edu_list)
                    except Exception:
                        edu_list = []

                edu_entries = []

                if isinstance(edu_list, list):

                    for edu in edu_list:

                        if not isinstance(edu, dict):
                            continue

                        school = (
                            edu.get("school_name")
                            or edu.get("schoolName")
                            or edu.get("school")
                            or ""
                        )

                        degree = edu.get("degree") or edu.get("degreeName") or ""

                        field = (
                            edu.get("field_of_study") or edu.get("fieldOfStudy") or ""
                        )

                        # Exclude IIIT entries from higher education
                        if school and "iiit" in school.lower():
                            continue

                        entry = school

                        if degree:
                            entry += f" ({degree})"

                        if field:
                            entry += f" - {field}"

                        if entry.strip():
                            edu_entries.append(entry)

                latest_higher_edu = "; ".join(edu_entries)

            if alumni.position_groups:
                position_groups_val = format_field(
                    alumni.position_groups, "position_groups"
                )

            # Normalization for comparison
            old_comp_norm = normalize_string(old_company)
            latest_comp_norm = normalize_string(db_company)

            old_des_norm = normalize_string(old_designation)
            latest_des_norm = normalize_string(db_designation)

            old_loc_norm = normalize_string(old_location)
            latest_loc_norm = normalize_string(db_location)

            old_edu_norm = normalize_string(old_higher_edu)
            latest_edu_norm = normalize_string(latest_higher_edu)

            # Detect changes (mismatch)
            if latest_comp_norm and old_comp_norm != latest_comp_norm:
                company_changed = True
            if latest_des_norm and old_des_norm != latest_des_norm:
                designation_changed = True
            if latest_loc_norm and old_loc_norm != latest_loc_norm:
                location_changed = True
            if latest_edu_norm and old_edu_norm != latest_edu_norm:
                higher_edu_changed = True

        new_vals_dict = {
            "New Company": db_company if company_changed else "",
            "New Designation": db_designation if designation_changed else "",
            "New Higher Education": latest_higher_edu if higher_edu_changed else "",
            "Updated Location": db_location if location_changed else "",
            "Changed": (
                "True"
                if (
                    company_changed
                    or designation_changed
                    or location_changed
                    or higher_edu_changed
                )
                else ""
            ),
            "Position Groups": position_groups_val,
        }

        for col_name in actual_new_cols:
            row_values.append(new_vals_dict[col_name])

        ws.append(row_values)

        current_row = ws.max_row
        start_new_col_idx = len(orig_cols) + 1

        for offset, col_name in enumerate(actual_new_cols):
            cell_val = new_vals_dict[col_name]
            col_idx = start_new_col_idx + offset

            should_color = False
            if col_name == "New Company" and company_changed:
                should_color = True
            elif col_name == "New Designation" and designation_changed:
                should_color = True
            elif col_name == "New Higher Education" and higher_edu_changed:
                should_color = True
            elif col_name == "Updated Location" and location_changed:
                should_color = True
            elif col_name == "Changed" and cell_val == "True":
                should_color = True

            if should_color:
                ws.cell(row=current_row, column=col_idx).fill = yellow_fill

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    out_filename = file_record.filename
    if not out_filename.lower().endswith(".xlsx"):
        out_filename = os.path.splitext(out_filename)[0] + ".xlsx"

    response = StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response.headers["Content-Disposition"] = f"attachment; filename={out_filename}"

    return response


@router.get("/download-original/{file_id}")
def download_original_file(file_id: int, db: Session = Depends(get_db)):
    file_record = db.query(UploadedFile).filter(UploadedFile.id == file_id).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="File history not found")

    file_path = os.path.join(UPLOAD_DIR, f"{file_record.id}_{file_record.filename}")
    if not os.path.exists(file_path):
        raise HTTPException(
            status_code=404, detail="Physical file no longer exists on the server"
        )

    return FileResponse(
        path=file_path, filename=file_record.filename, media_type="text/csv"
    )
