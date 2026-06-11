import os
import pandas as pd
import openpyxl
import subprocess

def create_mock_input():
    data = {
        "Work Location": [
            "NYC, USA",                      # Cities Match, Countries Match (New York, United States)
            "Bangalore, India",              # Cities Mismatch (Bangalore vs Chennai), Countries Match (India)
            "London, UK",                    # Cities Match (London), Countries Mismatch (United Kingdom vs Canada)
            "Munich, Germany",               # Cities Match, Countries Match (Munich, Germany)
            "Paris"                          # Empty / Missing Updated
        ],
        "Updated Location": [
            "New York City, United States",
            "Chennai, India",
            "London, Canada",
            "München, Deutschland",
            ""
        ]
    }
    df = pd.DataFrame(data)
    df.to_excel("input.xlsx", index=False)
    print("Created mock input.xlsx successfully.")

def run_matcher():
    print("Running location_matcher.py...")
    # Run using the python executable in the venv
    python_path = os.path.join("backend", "venv", "Scripts", "python.exe")
    if not os.path.exists(python_path):
        python_path = os.path.join("backend", "venv", "bin", "python")
    if not os.path.exists(python_path):
        python_path = "python" # fallback
        
    result = subprocess.run([python_path, "location_matcher.py"], capture_output=True, text=True)
    print("STDOUT:")
    print(result.stdout)
    print("STDERR:")
    print(result.stderr)
    assert result.returncode == 0, f"Matcher script failed with exit code {result.returncode}"

def verify_output():
    print("Verifying output.xlsx...")
    assert os.path.exists("output.xlsx"), "output.xlsx was not created"
    
    wb = openpyxl.load_workbook("output.xlsx")
    ws = wb.active
    
    # Map headers to column indices
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    
    work_city_idx = header_map.get('Work City')
    work_country_idx = header_map.get('Work Country')
    updated_city_idx = header_map.get('Updated City')
    updated_country_idx = header_map.get('Updated Country')
    
    assert work_city_idx and work_country_idx and updated_city_idx and updated_country_idx, "Output columns missing"
    
    # Helper to check if cell is filled with yellow (FFFF00 or 00FFFF00)
    def is_yellow(cell):
        fill = cell.fill
        if not fill or not fill.fill_type or fill.fill_type == 'none':
            return False
        rgb = fill.start_color.rgb if fill.start_color else None
        if not rgb:
            return False
        rgb_str = str(rgb).upper()
        return rgb_str in ('FFFF00', '00FFFF00', 'FFFFFF00', 'FFFFFFFF00') or rgb_str.endswith('FFFF00')

    # Row 2 (NYC vs New York City) -> Match -> no highlights
    r2_w_city = ws.cell(row=2, column=work_city_idx)
    r2_u_city = ws.cell(row=2, column=updated_city_idx)
    r2_w_country = ws.cell(row=2, column=work_country_idx)
    r2_u_country = ws.cell(row=2, column=updated_country_idx)
    
    print(f"Row 2 values: City='{r2_w_city.value}' vs '{r2_u_city.value}', Country='{r2_w_country.value}' vs '{r2_u_country.value}'")
    print(f"Row 2 fills: City={r2_w_city.fill.start_color.rgb if r2_w_city.fill else None}, Country={r2_w_country.fill.start_color.rgb if r2_w_country.fill else None}")
    assert not is_yellow(r2_w_city) and not is_yellow(r2_u_city), "Row 2 cities should match and not be highlighted"
    assert not is_yellow(r2_w_country) and not is_yellow(r2_u_country), "Row 2 countries should match and not be highlighted"
    
    # Row 3 (Bangalore vs Chennai) -> City mismatch -> highlighted city cells
    r3_w_city = ws.cell(row=3, column=work_city_idx)
    r3_u_city = ws.cell(row=3, column=updated_city_idx)
    r3_w_country = ws.cell(row=3, column=work_country_idx)
    r3_u_country = ws.cell(row=3, column=updated_country_idx)
    
    print(f"Row 3 values: City='{r3_w_city.value}' vs '{r3_u_city.value}', Country='{r3_w_country.value}' vs '{r3_u_country.value}'")
    assert is_yellow(r3_w_city) and is_yellow(r3_u_city), "Row 3 cities should mismatch and be highlighted"
    assert not is_yellow(r3_w_country) and not is_yellow(r3_u_country), "Row 3 countries should match and not be highlighted"

    # Row 4 (London, UK vs London, Canada) -> Country mismatch -> highlighted country cells
    r4_w_city = ws.cell(row=4, column=work_city_idx)
    r4_u_city = ws.cell(row=4, column=updated_city_idx)
    r4_w_country = ws.cell(row=4, column=work_country_idx)
    r4_u_country = ws.cell(row=4, column=updated_country_idx)
    
    print(f"Row 4 values: City='{r4_w_city.value}' vs '{r4_u_city.value}', Country='{r4_w_country.value}' vs '{r4_u_country.value}'")
    assert not is_yellow(r4_w_city) and not is_yellow(r4_u_city), "Row 4 cities should match and not be highlighted"
    assert is_yellow(r4_w_country) and is_yellow(r4_u_country), "Row 4 countries should mismatch and be highlighted"

    # Row 5 (Munich vs München, Germany vs Deutschland) -> Match -> no highlights
    r5_w_city = ws.cell(row=5, column=work_city_idx)
    r5_u_city = ws.cell(row=5, column=updated_city_idx)
    r5_w_country = ws.cell(row=5, column=work_country_idx)
    r5_u_country = ws.cell(row=5, column=updated_country_idx)
    
    print(f"Row 5 values: City='{r5_w_city.value}' vs '{r5_u_city.value}', Country='{r5_w_country.value}' vs '{r5_u_country.value}'")
    assert not is_yellow(r5_w_city) and not is_yellow(r5_u_city), "Row 5 cities should match and not be highlighted"
    assert not is_yellow(r5_w_country) and not is_yellow(r5_u_country), "Row 5 countries should match and not be highlighted"

    print("All assertions and verifications passed successfully!")

if __name__ == "__main__":
    create_mock_input()
    run_matcher()
    verify_output()
